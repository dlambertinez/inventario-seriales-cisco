#!/usr/bin/env python3

import argparse
import getpass
import ipaddress
import re
import sys
import traceback
from pathlib import Path
from typing import Optional

from netmiko import ConnectHandler
from netmiko.exceptions import (
    NetmikoAuthenticationException,
    NetmikoTimeoutException,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


NOMBRE_ARCHIVO_IPS = "ips.txt"
NOMBRE_ARCHIVO_EXCEL = "seriales_cisco.xlsx"
NOMBRE_ARCHIVO_ERROR = "error_programa.log"


def obtener_directorio_programa() -> Path:
    """
    Obtiene la carpeta en la que se encuentra el programa.

    Si se ejecuta como archivo EXE, retorna la carpeta del ejecutable.
    Si se ejecuta como archivo Python, retorna la carpeta del archivo .py.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent

    return Path(__file__).resolve().parent


DIRECTORIO_PROGRAMA = obtener_directorio_programa()


def resolver_ruta(nombre_archivo: str) -> Path:
    """
    Resuelve rutas relativas tomando como base la carpeta del programa.
    """
    ruta = Path(nombre_archivo)

    if ruta.is_absolute():
        return ruta

    return DIRECTORIO_PROGRAMA / ruta


def pausar_programa(
    mensaje: str = "\nPresiona Enter para cerrar el programa..."
) -> None:
    """
    Evita que la ventana se cierre automáticamente.
    """
    try:
        input(mensaje)
    except (EOFError, KeyboardInterrupt):
        pass


def leer_direcciones_ip(ruta_archivo: Path) -> list"""
    Lee las direcciones IP desde un archivo de texto.

    Se ignoran:
    - Líneas vacías.
    - Líneas que comienzan con #.
    - Texto ubicado después de #.
    - Direcciones IP duplicadas.
    """
    if not ruta_archivo.exists():
        raise FileNotFoundError(
            f"No se encontró el archivo de direcciones IP: "
            f"{ruta_archivo.resolve()}"
        )

    direcciones: list[str] = []
    direcciones_vistas: set[str] = set()

    with ruta_archivo.open("r", encoding="utf-8-sig") as archivo:
        for numero_linea, linea in enumerate(archivo, start=1):
            contenido = linea.split("#", maxsplit=1)[0].strip()

            if not contenido:
                continue

            try:
                ipaddress.ip_address(contenido)
            except ValueError:
                print(
                    f"[ADVERTENCIA] Línea {numero_linea} ignorada. "
                    f"No es una dirección IP válida: {contenido}"
                )
                continue

            if contenido not in direcciones_vistas:
                direcciones.append(contenido)
                direcciones_vistas.add(contenido)

    return direcciones


def valor_serial_valido(serial: Optional[str]) -> bool:
    """
    Verifica que el serial no sea vacío o un valor genérico.
    """
    if not serial:
        return False

    serial_normalizado = serial.strip().strip('"').upper()

    valores_no_validos = {
        "",
        "N/A",
        "NA",
        "NONE",
        "UNKNOWN",
        "NOT_SPECIFIED",
        "NOT SPECIFIED",
        "FFFFFFFFFFF",
    }

    return serial_normalizado not in valores_no_validos


def extraer_serial_show_inventory(salida: str) -> Optional"""
    Extrae el serial principal desde la salida de show inventory.
    """
    patron_bloque = re.compile(
        r'NAME:\s*"(?P<name>.*?)"\s*,\s*'
        r'DESCR:\s*"(?P<description>.*?)".*?'
        r"PID:\s*(?P<pid>[^,\r\n]*).*?"
        r"VID:\s*(?P<vid>[^,\r\n]*).*?"
        r"SN:\s*(?P<serial>[^\s,\r\n]+)",
        re.IGNORECASE | re.DOTALL,
    )

    inventario: list[dict[str, str]] = []

    for coincidencia in patron_bloque.finditer(salida):
        elemento = {
            "name": coincidencia.group("name").strip(),
            "description": coincidencia.group("description").strip(),
            "pid": coincidencia.group("pid").strip(),
            "serial": coincidencia.group("serial").strip().strip('"'),
        }

        if valor_serial_valido(elemento["serial"]):
            inventario.append(elemento)

    if not inventario:
        seriales = re.findall(
            r"\bSN:\s*([A-Za-z0-9._/-]+)",
            salida,
            flags=re.IGNORECASE,
        )

        for serial in seriales:
            if valor_serial_valido(serial):
                return serial.strip()

        return None

    palabras_equipo_principal = (
        "chassis",
        "router chassis",
        "switch chassis",
        "system",
    )

    for elemento in inventario:
        texto_elemento = (
            f"{elemento['name']} "
            f"{elemento['description']}"
        ).lower()

        if any(
            palabra in texto_elemento
            for palabra in palabras_equipo_principal
        ):
            return elemento["serial"]

    nombres_principales = {
        "chassis",
        "router",
        "switch 1",
        "switch 1 chassis",
    }

    for elemento in inventario:
        nombre = elemento["name"].strip().lower()

        if nombre in nombres_principales:
            return elemento["serial"]

    return inventario[0]["serial"]


def extraer_serial_show_version(salida: str) -> Optional"""
    Extrae el serial desde diferentes formatos de show version.
    """
    patrones = [
        r"Processor board ID\s+([A-Za-z0-9._/-]+)",
        r"System serial number\s*:\s*([A-Za-z0-9._/-]+)",
        r"System Serial Number\s*:\s*([A-Za-z0-9._/-]+)",
        r"Chassis Serial Number\s*:\s*([A-Za-z0-9._/-]+)",
        r"Motherboard serial number\s*:\s*([A-Za-z0-9._/-]+)",
        r"Serial Number\s*:\s*([A-Za-z0-9._/-]+)",
    ]

    for patron in patrones:
        coincidencia = re.search(
            patron,
            salida,
            flags=re.IGNORECASE,
        )

        if coincidencia:
            serial = coincidencia.group(1).strip()

            if valor_serial_valido(serial):
                return serial

    return None


def obtener_serial(
    direccion_ip: str,
    usuario: str,
    contrasena: str,
    secreto_enable: str = "",
) -> dict[str, str]:
    """
    Se conecta por SSH al dispositivo Cisco y obtiene su serial.
    """
    dispositivo = {
        "device_type": "cisco_ios",
        "host": direccion_ip,
        "username": usuario,
        "password": contrasena,
        "secret": secreto_enable,
        "port": 22,
        "conn_timeout": 15,
        "auth_timeout": 15,
        "banner_timeout": 20,
        "timeout": 30,
        "fast_cli": False,
    }

    conexion = None

    try:
        print(f"[CONECTANDO] {direccion_ip}")

        conexion = ConnectHandler(**dispositivo)

        if secreto_enable:
            try:
                conexion.enable()
                print(f"[ENABLE] {direccion_ip}: modo privilegiado activo")
            except Exception as error:
                mensaje = str(error).replace("\n", " ").strip()

                print(
                    f"[ERROR] {direccion_ip}: "
                    "no fue posible ingresar al modo enable"
                )

                return {
                    "ip": direccion_ip,
                    "serial": "",
                    "estado": (
                        "No fue posible ingresar al modo enable: "
                        f"{mensaje}"
                    ),
                }

        try:
            conexion.send_command(
                "terminal length 0",
                expect_string=r"#|>",
                read_timeout=15,
            )
        except Exception:
            pass

        salida_inventario = conexion.send_command(
            "show inventory",
            read_timeout=45,
        )

        serial = extraer_serial_show_inventory(salida_inventario)

        if serial:
            print(f"[CORRECTO] {direccion_ip} - Serial: {serial}")

            return {
                "ip": direccion_ip,
                "serial": serial,
                "estado": "Correcto - show inventory",
            }

        print(
            f"[INFORMACIÓN] {direccion_ip}: no se encontró el serial "
            "en show inventory. Consultando show version."
        )

        salida_version = conexion.send_command(
            "show version",
            read_timeout=45,
        )

        serial = extraer_serial_show_version(salida_version)

        if serial:
            print(f"[CORRECTO] {direccion_ip} - Serial: {serial}")

            return {
                "ip": direccion_ip,
                "serial": serial,
                "estado": "Correcto - show version",
            }

        print(f"[SIN SERIAL] {direccion_ip}")

        return {
            "ip": direccion_ip,
            "serial": "",
            "estado": (
                "Conexión correcta, pero no se encontró el serial"
            ),
        }

    except NetmikoAuthenticationException:
        print(
            f"[ERROR] {direccion_ip}: autenticación SSH incorrecta"
        )

        return {
            "ip": direccion_ip,
            "serial": "",
            "estado": "Error de autenticación SSH",
        }

    except NetmikoTimeoutException:
        print(
            f"[ERROR] {direccion_ip}: tiempo de conexión SSH agotado"
        )

        return {
            "ip": direccion_ip,
            "serial": "",
            "estado": "Tiempo de conexión SSH agotado",
        }

    except TimeoutError:
        print(
            f"[ERROR] {direccion_ip}: tiempo de conexión agotado"
        )

        return {
            "ip": direccion_ip,
            "serial": "",
            "estado": "Tiempo de conexión agotado",
        }

    except Exception as error:
        mensaje = str(error).replace("\n", " ").strip()

        print(f"[ERROR] {direccion_ip}: {mensaje}")

        return {
            "ip": direccion_ip,
            "serial": "",
            "estado": f"Error: {mensaje}",
        }

    finally:
        if conexion is not None:
            try:
                conexion.disconnect()
            except Exception:
                pass


def crear_archivo_excel(
    resultados: list[dict[str, str]],
    ruta_salida: Path,
) -> None:
    """
    Crea el archivo Excel con las columnas IP, Serial y Estado.
    """
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Seriales Cisco"

    encabezados = ["IP", "Serial", "Estado"]
    hoja.append(encabezados)

    color_encabezado = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    fuente_encabezado = Font(
        color="FFFFFF",
        bold=True,
    )

    for celda in hojacelda.fill = color_encabezado
        celda.font = fuente_encabezado
        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for resultado in resultados:
        hoja.append(
            [
                resultado["ip"],
                resultado["serial"],
                resultado["estado"],
            ]
        )

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    anchos_iniciales = {
        "A": 18,
        "B": 25,
        "C": 60,
    }

    for columna, ancho in anchos_iniciales.items():
        hoja.column_dimensions[columna].width = ancho

    for fila in hoja.iter_rows(min_row=2):
        fila[0].alignment = Alignment(vertical="top")
        fila[1].alignment = Alignment(vertical="top")
        fila[2].alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

    for columna in hoja.columns:
        numero_columna = columna[0].column
        letra_columna = get_column_letter(numero_columna)

        longitud_maxima = max(
            len(str(celda.value))
            if celda.value is not None
            else 0
            for celda in columna
        )

        ancho_actual = (
            hoja.column_dimensions[letra_columna].width or 10
        )

        ancho_calculado = min(longitud_maxima + 2, 70)

        hoja.column_dimensions[letra_columna].width = max(
            ancho_actual,
            ancho_calculado,
        )

    ruta_salida.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    libro.save(ruta_salida)


def obtener_argumentos() -> argparse.Namespace:
    """
    Obtiene argumentos opcionales indicados por línea de comandos.
    """
    parser = argparse.ArgumentParser(
        description=(
            "Obtiene por SSH los seriales de equipos Cisco "
            "y almacena los resultados en un archivo Excel."
        )
    )

    parser.add_argument(
        "-i",
        "--ips",
        default=NOMBRE_ARCHIVO_IPS,
        help=(
            "Archivo con una dirección IP por línea. "
            f"Predeterminado: {NOMBRE_ARCHIVO_IPS}"
        ),
    )

    parser.add_argument(
        "-o",
        "--output",
        default=NOMBRE_ARCHIVO_EXCEL,
        help=(
            "Nombre del archivo Excel de salida. "
            f"Predeterminado: {NOMBRE_ARCHIVO_EXCEL}"
        ),
    )

    return parser.parse_args()


def main() -> int:
    """
    Función principal del programa.
    """
    argumentos = obtener_argumentos()

    ruta_ips = resolver_ruta(argumentos.ips)
    ruta_excel = resolver_ruta(argumentos.output)

    print("=" * 70)
    print(" INVENTARIO DE SERIALES CISCO POR SSH")
    print("=" * 70)
    print(f"Carpeta del programa : {DIRECTORIO_PROGRAMA}")
    print(f"Archivo de IP        : {ruta_ips}")
    print(f"Archivo de resultados: {ruta_excel}")
    print("=" * 70)

    try:
        direcciones_ip = leer_direcciones_ip(ruta_ips)

    except FileNotFoundError as error:
        print(f"\n[ERROR] {error}")
        print(
            "\nDebes colocar el archivo ips.txt en la misma "
            "carpeta del ejecutable."
        )
        return 1

    except OSError as error:
        print(
            f"\n[ERROR] No fue posible leer el archivo "
            f"de direcciones IP: {error}"
        )
        return 1

    if not direcciones_ip:
        print(
            f"\n[ERROR] El archivo {ruta_ips} no contiene "
            "direcciones IP válidas."
        )
        return 1

    print(
        f"\nSe encontraron {len(direcciones_ip)} "
        "direcciones IP válidas.\n"
    )

    usuario = input("Usuario SSH: ").strip()

    if not usuario:
        print("\n[ERROR] El usuario SSH no puede estar vacío.")
        return 1

    contrasena = getpass.getpass("Contraseña SSH: ")

    if not contrasena:
        print("\n[ERROR] La contraseña SSH no puede estar vacía.")
        return 1

    secreto_enable = getpass.getpass(
        "Contraseña enable, opcional; "
        "presiona Enter para omitir: "
    )

    resultados: list[dict[str, str]] = []

    total_equipos = len(direcciones_ip)

    for posicion, direccion_ip in enumerate(
        direcciones_ip,
        start=1,
    ):
        print()
        print("-" * 70)
        print(
            f"Equipo {posicion} de {total_equipos}: "
            f"{direccion_ip}"
        )
        print("-" * 70)

        resultado = obtener_serial(
            direccion_ip=direccion_ip,
            usuario=usuario,
            contrasena=contrasena,
            secreto_enable=secreto_enable,
        )

        resultados.append(resultado)

    try:
        crear_archivo_excel(
            resultados=resultados,
            ruta_salida=ruta_excel,
        )

    except PermissionError:
        print(
            f"\n[ERROR] No se pudo guardar el archivo:\n"
            f"{ruta_excel}"
        )
        print(
            "\nVerifica que seriales_cisco.xlsx "
            "no esté abierto en Excel."
        )
        return 1

    except OSError as error:
        print(f"\n[ERROR] No se pudo crear el Excel: {error}")
        return 1

    correctos = sum(
        1
        for resultado in resultados
        if resultado["serial"]
    )

    fallidos = len(resultados) - correctos

    print()
    print("=" * 70)
    print(" PROCESO FINALIZADO")
    print("=" * 70)
    print(f"Equipos procesados : {len(resultados)}")
    print(f"Seriales obtenidos : {correctos}")
    print(f"Sin serial o error : {fallidos}")
    print(f"Archivo generado   : {ruta_excel.resolve()}")
    print("=" * 70)

    return 0


def ejecutar_programa() -> int:
    """
    Ejecuta el programa y registra cualquier error inesperado.
    """
    try:
        return main()

    except KeyboardInterrupt:
        print("\n\n[INFORMACIÓN] Proceso cancelado por el usuario.")
        return 130

    except Exception:
        detalle_error = traceback.format_exc()
        ruta_error = DIRECTORIO_PROGRAMA / NOMBRE_ARCHIVO_ERROR

        try:
            ruta_error.write_text(
                detalle_error,
                encoding="utf-8",
            )
        except OSError:
            pass

        print()
        print("=" * 70)
        print(" ERROR INESPERADO")
        print("=" * 70)
        print(detalle_error)
        print(
            "El detalle del error se guardó en:\n"
            f"{ruta_error}"
        )

        return 1


if __name__ == "__main__":
    codigo_salida = ejecutar_programa()
    pausar_programa()
    sys.exit(codigo_salida)
